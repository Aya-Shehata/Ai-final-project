from datetime import datetime
from openpyxl import load_workbook
from classes import City, Flight

workbook = load_workbook(filename="C:/Users/Dell/Downloads/Atrificial intelligence/Travel Agent KB (2 sheets).xlsx",read_only=True)
sheet = workbook.active

cities = [] # a list the has all the cities from excel sheet


for row in sheet.iter_rows(min_row=2, values_only=True):
    city = City(name=row[0],
                      latitude=row[1],
                      longitude=row[2])
    cities.append(city)

workbook = load_workbook(filename="C:/Users/Dell/Downloads/Atrificial intelligence/Travel Agent KB.xlsx",read_only=True)
sheet = workbook.active
listofflights = [] #a list has all the flights from excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    s1:City
    g1:City
    for i in range(len(cities)): #loop to create a new object of city class for each origin and destination cities
        if(cities[i].name == row[0]):
            s1 = cities[i]
        elif(cities[i].name == row[1]):
            g1 = cities[i]

    flight = Flight(s=s1,
                      d=g1,
                      dt=row[2],
                      at=row[3],
                      fn=row[4],
                      l=row[5])
    listofflights.append(flight)

#print(flights[0].departure_time)